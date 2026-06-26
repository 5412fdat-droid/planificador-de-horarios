import os
import django
import sys
import openpyxl

# Configurar Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from core.models import Carrera
from administrador.models import Materia, Docente, Aula, GrupoMateria, Horario
from django.db import transaction

def color_to_hex(color_index):
    if not color_index or color_index == '00000000':
        return '#3b82f6' # default blue
    
    if type(color_index) is int or color_index.isdigit():
        # It's a theme index (like 3)
        return '#3b82f6'
    
    # It's an ARGB hex string like 'FFFF0000'
    if len(color_index) == 8:
        return '#' + color_index[2:]
    return '#' + color_index

def parse_time(time_str):
    if not time_str:
        return None, None
    time_str = time_str.lower().replace(' a ', '-').replace(' - ', '-').strip()
    parts = time_str.split('-')
    if len(parts) == 2:
        return parts[0].strip() + ':00', parts[1].strip() + ':00'
    return None, None

def main():
    file_path = r"c:\Users\Nuevo Usuario\Documents\proyecto de sistemas de informacion II\proyecto\HORARIOS  ING. EN SISTEMAS  I-2026.xlsx"
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    carrera = Carrera.objects.filter(nombre_carrera__icontains='SISTEMAS').first()
    if not carrera:
        print("Carrera Ingeniería en Sistemas no encontrada.")
        return

    # Mapeos temporales
    dias = ['Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes']
    
    # Limpiar horarios existentes para Ingeniería en Sistemas en la gestión 1-2026 (solo los que vamos a cargar)
    # Mejor no borrar ciegamente, solo si es necesario. Como es inserción limpia de semestres impares, 
    # borramos los grupos de los semestres 1,3,5,7,9 de Sistemas gestión 1-2026
    grupos_a_borrar = GrupoMateria.objects.filter(
        materia__carrera=carrera,
        materia__periodo_materia__in=['1', '3', '5', '7', '9'],
        gestion='1-2026'
    )
    grupos_a_borrar.delete()

    print("Iniciando importación...")

    current_semestre = None
    current_aula = None
    
    # 1. PRE-READ LEGEND
    print("Pre-leyendo leyenda...")
    docente_materia_map = {}
    for row in range(1, sheet.max_row + 1):
        docente_val = sheet.cell(row=row, column=9).value
        sigla_val = sheet.cell(row=row, column=10).value
        if sigla_val and isinstance(sigla_val, str) and sigla_val.upper() != 'SIGLA':
            docente_name = str(docente_val).strip() if docente_val else 'Por Asignar'
            sigla_materia = sigla_val.strip().replace(' ', '-')
            if len(sigla_materia.split('-')) > 1:
                part1, part2 = sigla_materia.split('-')
                sigla_materia = f"{part1.strip()}-{part2.strip()}"
            
            docente = Docente.objects.filter(nombre_docente__icontains=docente_name[:10]).first()
            if not docente:
                import uuid
                codigo = (docente_name[:10].replace(' ', '').upper() + str(uuid.uuid4().hex[:4]))
                docente = Docente.objects.create(nombre_docente=docente_name, codigo_docente=codigo)
            
            materia = Materia.objects.filter(sigla_materia=sigla_materia, carrera=carrera).first()
            if materia:
                cell_color = str(sheet.cell(row=row, column=10).fill.start_color.index)
                color_hex = color_to_hex(cell_color)
                docente_materia_map[sigla_materia] = {
                    'materia': materia,
                    'docente': docente,
                    'color_hex': color_hex
                }

    # Helper function for merged cells
    def get_merged_range(row, col):
        for merged_cell in sheet.merged_cells.ranges:
            if merged_cell.min_row <= row <= merged_cell.max_row and merged_cell.min_col <= col <= merged_cell.max_col:
                return merged_cell
        return None

    # 2. READ GRID
    print("Leyendo grilla de horarios...")
    for row in range(1, sheet.max_row + 1):
        cell_semestre = sheet.cell(row=row, column=2).value
        if cell_semestre and isinstance(cell_semestre, str) and 'SEMESTRE' in cell_semestre.upper() and 'OFERTA' not in cell_semestre.upper() and 'CARRERA' not in cell_semestre.upper():
            s_num = ''.join([c for c in cell_semestre if c.isdigit()])
            if s_num in ['1', '3', '5', '7', '9']:
                current_semestre = s_num
                print(f"-> Procesando Semestre {current_semestre}")
            else:
                current_semestre = None
        
        if current_semestre:
            cell_aula = sheet.cell(row=row, column=2).value
            if cell_aula and isinstance(cell_aula, str) and 'AULA' in cell_aula.upper():
                aula_name = cell_aula.split(':')[-1].strip()
                if aula_name:
                    aula, _ = Aula.objects.get_or_create(nombre_aula=f"Aula {aula_name}", defaults={'capacidad_maxima': 50})
                    current_aula = aula

            # Procesar Horario (Columnas 2 a 7)
            time_val = sheet.cell(row=row, column=2).value
            if time_val and isinstance(time_val, str) and ('a' in time_val or '-' in time_val) and ':' in time_val:
                # We need to process each column. If the cell is the TOP of a merged cell, or an unmerged cell with value
                for col_idx, dia in enumerate(dias, start=3):
                    celda_dia = sheet.cell(row=row, column=col_idx)
                    val = celda_dia.value
                    
                    merged = get_merged_range(row, col_idx)
                    if merged:
                        # Only process the top-left cell of the merged range to avoid duplicates
                        if row != merged.min_row or col_idx != merged.min_col:
                            continue
                        val = sheet.cell(merged.min_row, merged.min_col).value
                        # Calculate time from min_row to max_row
                        time_start_str = sheet.cell(merged.min_row, 2).value
                        time_end_str = sheet.cell(merged.max_row, 2).value
                        h_ini, _ = parse_time(str(time_start_str)) if time_start_str else (None, None)
                        _, h_fin = parse_time(str(time_end_str)) if time_end_str else (None, None)
                    else:
                        h_ini, h_fin = parse_time(time_val)

                    if val and isinstance(val, str) and h_ini and h_fin:
                        sigla_bloque = val.strip().replace(' ', '-')
                        if len(sigla_bloque.split('-')) > 1:
                            part1, part2 = sigla_bloque.split('-')
                            sigla_bloque = f"{part1.strip()}-{part2.strip()}"
                        
                        c_index = str(celda_dia.fill.start_color.index) if celda_dia.fill else None
                        bloque_color = color_to_hex(c_index)
                        
                        info = docente_materia_map.get(sigla_bloque)
                        if not info:
                            mat = Materia.objects.filter(sigla_materia__icontains=sigla_bloque, carrera=carrera).first()
                            if mat:
                                doc, _ = Docente.objects.get_or_create(nombre_docente='Por Asignar', defaults={'codigo_docente': 'PORASIGNAR'})
                                info = {'materia': mat, 'docente': doc, 'color_hex': bloque_color}
                        
                        if info:
                            with transaction.atomic():
                                grupo, _ = GrupoMateria.objects.get_or_create(
                                    materia=info['materia'],
                                    docente=info['docente'],
                                    gestion='1-2026',
                                    nombre_grupo='AV' if current_semestre in ['1', '3', '5'] else 'VA',
                                    defaults={
                                        'cupo_limite': 50,
                                        'color_hex': info['color_hex']
                                    }
                                )
                                if grupo.color_hex != info['color_hex']:
                                    grupo.color_hex = info['color_hex']
                                    grupo.save()
                                
                                Horario.objects.create(
                                    grupo_materia=grupo,
                                    aula=current_aula,
                                    dia_semana=dia,
                                    hora_inicio=h_ini,
                                    hora_fin=h_fin
                                )
                                print(f"  + {dia} {h_ini}-{h_fin}: {sigla_bloque} en {current_aula.nombre_aula if current_aula else 'S/A'} (Docente: {info['docente'].nombre_docente})")

    print("Importación completada con éxito.")

if __name__ == '__main__':
    main()
