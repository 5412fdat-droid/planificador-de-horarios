import openpyxl

file_path = r"c:\Users\Nuevo Usuario\Documents\proyecto de sistemas de informacion II\proyecto\HORARIOS  ING. EN SISTEMAS  I-2026.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active

for merged_cell in sheet.merged_cells.ranges:
    # merged_cell.min_row, merged_cell.max_row, merged_cell.min_col, merged_cell.max_col
    if merged_cell.min_row >= 7 and merged_cell.min_row <= 25 and merged_cell.min_col >= 3 and merged_cell.max_col <= 7:
        print(f"Merged: row {merged_cell.min_row}-{merged_cell.max_row}, col {merged_cell.min_col}-{merged_cell.max_col}, val: {sheet.cell(merged_cell.min_row, merged_cell.min_col).value}")

