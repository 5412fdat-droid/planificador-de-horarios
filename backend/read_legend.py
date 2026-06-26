import openpyxl

file_path = r"c:\Users\Nuevo Usuario\Documents\proyecto de sistemas de informacion II\proyecto\HORARIOS  ING. EN SISTEMAS  I-2026.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active

for merged_cell in sheet.merged_cells.ranges:
    # check if in legend (col 9, 10, 11)
    if merged_cell.min_col >= 9 and merged_cell.max_col <= 11 and merged_cell.min_row >= 7 and merged_cell.min_row <= 25:
        print(f"Legend Merged: row {merged_cell.min_row}-{merged_cell.max_row}, col {merged_cell.min_col}-{merged_cell.max_col}, val: {sheet.cell(merged_cell.min_row, merged_cell.min_col).value}")

for row in range(8, 16):
    doc = sheet.cell(row, 9).value
    sigla = sheet.cell(row, 10).value
    c_index = sheet.cell(row, 10).fill.start_color.index
    print(f"Row {row}: {doc} | {sigla} | color: {c_index}")
