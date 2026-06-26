import openpyxl

file_path = r"c:\Users\Nuevo Usuario\Documents\proyecto de sistemas de informacion II\proyecto\HORARIO   INDUSTRIALIZACIÓN DE ALIMENTOS I-2026.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"--- Sheet: {sheet_name} ---")
    
    # Let's print out rows that look like a legend
    for row in range(1, 40):
        row_data = []
        for col in range(1, 15):
            val = sheet.cell(row=row, column=col).value
            if val is not None:
                row_data.append(str(val))
        if len(row_data) > 0:
            print(f"Row {row}: {row_data}")
